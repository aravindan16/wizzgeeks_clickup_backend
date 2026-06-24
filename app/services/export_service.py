"""Render a normalized report (columns + rows + summary) to Excel or PDF bytes."""
from io import BytesIO
from typing import Any

from openpyxl import Workbook
from openpyxl.styles import Alignment, Font, PatternFill
from openpyxl.utils import get_column_letter
from reportlab.lib import colors
from reportlab.lib.pagesizes import A4, landscape
from reportlab.lib.styles import getSampleStyleSheet
from reportlab.lib.units import mm
from reportlab.platypus import Paragraph, SimpleDocTemplate, Spacer, Table, TableStyle

HEADER_FILL = PatternFill(start_color="111827", end_color="111827", fill_type="solid")
HEADER_FONT = Font(color="FFFFFF", bold=True)


def to_excel(report: dict[str, Any]) -> bytes:
    wb = Workbook()
    ws = wb.active
    ws.title = "Report"

    columns = report["columns"]
    rows = report["rows"]

    # Title.
    ws.append([report["title"]])
    ws["A1"].font = Font(size=14, bold=True)
    ws.append([f"Generated: {report['generated_at']}"])
    ws.append([])

    header_row_idx = ws.max_row + 1
    ws.append([c["label"] for c in columns])
    for col_idx in range(1, len(columns) + 1):
        cell = ws.cell(row=header_row_idx, column=col_idx)
        cell.fill = HEADER_FILL
        cell.font = HEADER_FONT
        cell.alignment = Alignment(horizontal="center")

    for r in rows:
        ws.append([r.get(c["key"], "") for c in columns])

    # Auto-ish width.
    for col_idx, c in enumerate(columns, start=1):
        values = [str(c["label"])] + [str(r.get(c["key"], "")) for r in rows]
        width = min(50, max(10, max(len(v) for v in values) + 2))
        ws.column_dimensions[get_column_letter(col_idx)].width = width

    # Summary block.
    ws.append([])
    ws.append(["Summary"])
    ws.cell(row=ws.max_row, column=1).font = Font(bold=True)
    for k, v in report["summary"].items():
        ws.append([k, str(v)])

    buf = BytesIO()
    wb.save(buf)
    return buf.getvalue()


def to_pdf(report: dict[str, Any]) -> bytes:
    buf = BytesIO()
    doc = SimpleDocTemplate(buf, pagesize=landscape(A4),
                            leftMargin=12 * mm, rightMargin=12 * mm,
                            topMargin=12 * mm, bottomMargin=12 * mm)
    styles = getSampleStyleSheet()
    elements: list = [Paragraph(report["title"], styles["Title"])]
    elements.append(Paragraph(f"Generated: {report['generated_at']}", styles["Normal"]))
    elements.append(Spacer(1, 8))

    columns = report["columns"]
    header = [c["label"] for c in columns]
    data = [header]
    for r in report["rows"]:
        data.append([str(r.get(c["key"], "")) for c in columns])

    if len(data) == 1:
        data.append(["No data"] + [""] * (len(columns) - 1))

    table = Table(data, repeatRows=1)
    table.setStyle(TableStyle([
        ("BACKGROUND", (0, 0), (-1, 0), colors.HexColor("#111827")),
        ("TEXTCOLOR", (0, 0), (-1, 0), colors.white),
        ("FONTSIZE", (0, 0), (-1, -1), 7),
        ("GRID", (0, 0), (-1, -1), 0.25, colors.HexColor("#cccccc")),
        ("ROWBACKGROUNDS", (0, 1), (-1, -1), [colors.white, colors.HexColor("#f3f4f6")]),
        ("VALIGN", (0, 0), (-1, -1), "TOP"),
    ]))
    elements.append(table)

    elements.append(Spacer(1, 12))
    elements.append(Paragraph("Summary", styles["Heading3"]))
    summary_data = [[k, str(v)] for k, v in report["summary"].items()]
    if summary_data:
        st = Table(summary_data, colWidths=[60 * mm, 80 * mm])
        st.setStyle(TableStyle([
            ("FONTSIZE", (0, 0), (-1, -1), 8),
            ("GRID", (0, 0), (-1, -1), 0.25, colors.HexColor("#cccccc")),
            ("BACKGROUND", (0, 0), (0, -1), colors.HexColor("#f3f4f6")),
        ]))
        elements.append(st)

    doc.build(elements)
    return buf.getvalue()
